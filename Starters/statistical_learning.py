from pybaseball import *
import pandas as pd
import math
from statistics import mean
import matplotlib.pyplot as plt
from sklearn.preprocessing import PolynomialFeatures
from sklearn.linear_model import LinearRegression, Lasso, Ridge, ElasticNet
from sklearn.metrics import mean_squared_error, r2_score
from sklearn.model_selection import train_test_split
import numpy as np
import pickle
from pathlib import Path
from openpyxl import load_workbook

'''
      batting_stats index:
      Season, Name, Team: 1~3
      G, AB, PA, H: 5~8, 2B: 10, HR: 12, AVG: 24, BB%: 35, K%: 36
      OBP, SLG, OPS, ISO, BABIP: 38~42
      LD%, GB%, FB%: 44~46
      
      Others:
      a = stats[0:15] is a Pandas Dataframe
      to_numpy() converts into a 2D Numpy array
'''

base_dir = Path(__file__).resolve().parent
projections = ['ATC', 'THE_BAT', 'Steamer', 'ZIPS', 'ZIPS_DC', 'DC', 'OOPSY']

def compute_rmse_by_stat(df, stats):
    rmse = {}
    for stat in stats:
        real_col = f'real_{stat}'
        if stat not in df.columns or real_col not in df.columns:
            rmse[stat] = None
            continue
        pair = df[[stat, real_col]].apply(pd.to_numeric, errors='coerce').dropna()
        if pair.empty:
            rmse[stat] = None
            continue
        rmse[stat] = float(np.sqrt(mean_squared_error(pair[real_col], pair[stat])))
    return rmse


def collect_2025_rmse():
    rmse_table = {}
    stat_order = None
    for proj in projections:
        path = base_dir / proj / f'2025_{proj}_Projections.csv'
        if not path.exists():
            print(f'{proj}: missing file {path}, skipping')
            continue

        df = pd.read_csv(path)

        stats = [c for c in df.columns if c not in {'#', 'Name', 'Team'} and not c.startswith('real_')]
        if stat_order is None:
            stat_order = stats
        rmse_table[proj] = compute_rmse_by_stat(df, stats)
    return rmse_table, (stat_order or [])


def clear_sheet(ws):
    for r in range(1, ws.max_row + 1):
        for c in range(1, ws.max_column + 1):
            ws.cell(r, c).value = None


def write_model_comparisons_2025(rmse_table, stat_order):
    workbook_path = base_dir / 'Model_Comparisons.xlsx'
    wb = load_workbook(workbook_path)
    ws = wb['Sheet1']
    clear_sheet(ws)

    ws.cell(1, 1).value = 'Model Performance (Metric: RMSE)'

    display_models = [
        ('ATC', 'ATC'),
        ('THE BAT', 'THE_BAT'),
        ('Steamer', 'Steamer'),
        ('ZIPS', 'ZIPS'),
        ('ZIPS DC', 'ZIPS_DC'),
        ('DC', 'DC'),
        ('OOPSY', 'OOPSY')
    ]

    # Keep existing workbook style: 3 stat groups per section, each group width = 8 cols.
    row = 2
    for i in range(0, len(stat_order), 3):
        for group in range(3):
            stat_idx = i + group
            col0 = 1 + group * 8
            if stat_idx >= len(stat_order):
                continue

            stat = stat_order[stat_idx]
            ws.cell(row, col0).value = stat
            ws.cell(row + 1, col0).value = None
            ws.cell(row + 2, col0).value = 2025
            ws.cell(row + 3, col0).value = None
            ws.cell(row + 4, col0).value = None

            for j, (display_name, model_key) in enumerate(display_models, start=1):
                ws.cell(row + 1, col0 + j).value = display_name
                ws.cell(row + 2, col0 + j).value = rmse_table.get(model_key, {}).get(stat)
                ws.cell(row + 3, col0 + j).value = None
                ws.cell(row + 4, col0 + j).value = None

        row += 5

    wb.save(workbook_path)
    print(f'Updated workbook: {workbook_path}')


def print_rmse_results_2025(rmse_table, stat_order):
    model_order = ['ATC', 'THE_BAT', 'Steamer', 'ZIPS', 'ZIPS_DC', 'DC', 'OOPSY']
    print("\n2025 RMSE Results")
    for stat in stat_order:
        print(f"\n{stat}")
        for model in model_order:
            value = rmse_table.get(model, {}).get(stat)
            if value is None:
                print(f"  {model}: N/A")
            else:
                print(f"  {model}: {value:.6f}")


rmse_table, stats_2025 = collect_2025_rmse()
print_rmse_results_2025(rmse_table, stats_2025)
# write_model_comparisons_2025(rmse_table, stats_2025)

# print(find_player_name('guerrero jr.','vladimir'))
# é
