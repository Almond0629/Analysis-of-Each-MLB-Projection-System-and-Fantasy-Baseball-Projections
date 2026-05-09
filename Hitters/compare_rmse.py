from pybaseball import *
import pandas as pd
from sklearn.metrics import mean_squared_error, r2_score
import numpy as np
from pathlib import Path
from openpyxl import load_workbook

'''
      batting_stats index:
      Season, Name, Team: 1~3
      G, AB, PA, H: 5~8, 2B: 10, HR: 12, AVG: 24, BB%: 35, K%: 36
      OBP, SLG, OPS, ISO, BABIP: 38~42
      LD%, GB%, FB%: 44~46
      
      Others:
      a = stats[0:15] is a Pandas Dataframe
      to_numpy() converts into a 2D Numpy array
'''

def add_derived_stat_columns(df):
    df = df.copy()
    df['HIP'] = df['H'] - df['HR']
    df['real_HIP'] = df['real_H'] - df['real_HR']

    df['1B%'] = df['1B'] / df['HIP'].replace(0, np.nan)
    df['real_1B%'] = df['real_1B'] / df['real_HIP'].replace(0, np.nan)
    df['2B%'] = df['2B'] / df['HIP'].replace(0, np.nan)
    df['real_2B%'] = df['real_2B'] / df['real_HIP'].replace(0, np.nan)
    df['3B%'] = df['3B'] / df['HIP'].replace(0, np.nan)
    df['real_3B%'] = df['real_3B'] / df['real_HIP'].replace(0, np.nan)
    df['HR%'] = df['HR'] / df['PA'].replace(0, np.nan)
    df['real_HR%'] = df['real_HR'] / df['real_PA'].replace(0, np.nan)

    if 'NSB' not in df.columns and {'SB', 'CS'}.issubset(df.columns):
        df['NSB'] = df['SB'] - df['CS']
    if 'real_NSB' not in df.columns and {'real_SB', 'real_CS'}.issubset(df.columns):
        df['real_NSB'] = df['real_SB'] - df['real_CS']

    return df


def print_rmse_for_stats(df, label):
    rmse_stats = ['H', '1B%', '2B%', '3B%', 'HR%', 'R', 'RBI', 'BB', 'IBB', 'SO', 'HBP', 'SF', 'SH', 'SB', 'CS', 'NSB', 'AVG', 'OBP', 'SLG', 'OPS', 'BABIP', 'K%', 'BB%']
    print(f'  {label}')
    for stat_name in rmse_stats:
        real_col = f'real_{stat_name}'
        if stat_name not in df.columns or real_col not in df.columns:
            print(f'    {stat_name}: missing column(s)')
            continue

        pair_df = df[[stat_name, real_col]].apply(pd.to_numeric, errors='coerce').dropna()
        if pair_df.empty:
            print(f'    {stat_name}: no valid rows')
            continue

        rmse = np.sqrt(mean_squared_error(pair_df[real_col], pair_df[stat_name]))
        print(f'    {stat_name}: {rmse:.6f}')


projections = ['ATC','THE_BAT','THE_BAT_X','Steamer','ZIPS','ZIPS_DC','DC','OOPSY']
base_dir = Path(__file__).resolve().parent
excluded_stats_for_pct = {'#', 'Name', 'Team', 'G', 'AB', 'PA'}


def compute_rmse_by_stat(df, stats):
    result = {}
    for stat_name in stats:
        real_col = f'real_{stat_name}'
        if stat_name not in df.columns or real_col not in df.columns:
            result[stat_name] = None
            continue

        pair_df = df[[stat_name, real_col]].apply(pd.to_numeric, errors='coerce').dropna()
        if pair_df.empty:
            result[stat_name] = None
            continue

        result[stat_name] = float(np.sqrt(mean_squared_error(pair_df[real_col], pair_df[stat_name])))
    return result

def get_percentage_stats(df):
    stats = []
    for col in df.columns:
        if col.startswith('real_'):
            continue
        if col in excluded_stats_for_pct:
            continue
        if col.startswith('Unnamed'):
            continue
        if f'real_{col}' not in df.columns:
            continue
        stats.append(col)
    return stats

def collect_percentage_rmse_2023_2025(years=(2023, 2024, 2025), projection_systems=None):
    if projection_systems is None:
        projection_systems = projections

    rmse_table = {}
    for proj in projection_systems:
        rmse_table[proj] = {}
        yearly_frames = []
        for year in years:
            path = base_dir / proj / f'{year}_{proj}_Projections.csv'
            if not path.exists():
                rmse_table[proj][year] = None
                continue

            df = pd.read_csv(path)
            yearly_frames.append(df)
            stats = get_percentage_stats(df)
            rmse_table[proj][year] = compute_rmse_by_stat(df, stats)

        if yearly_frames:
            combined_df = pd.concat(yearly_frames, ignore_index=True, sort=False)
            combined_stats = get_percentage_stats(combined_df)
            rmse_table[proj]['Total'] = compute_rmse_by_stat(combined_df, combined_stats)
        else:
            rmse_table[proj]['Total'] = None
    return rmse_table

def print_percentage_rmse_2023_2025(rmse_table, years=(2023, 2024, 2025), projection_systems=None):
    if projection_systems is None:
        projection_systems = projections

    all_stats = set()
    for proj in projection_systems:
        for period in ['Total', *years]:
            period_result = rmse_table.get(proj, {}).get(period)
            if isinstance(period_result, dict):
                all_stats.update(period_result.keys())

    model_header = "     " + "   ".join(projection_systems)
    row_periods = ['Total', 2025, 2024, 2023]

    print('\nRMSE of PA-Based Stat Percentages (real_stat as ground truth)')
    for stat in sorted(all_stats):
        print(f'\n{stat}')
        print(model_header)
        for period in row_periods:
            values = []
            for proj in projection_systems:
                period_result = rmse_table.get(proj, {}).get(period)
                value = None
                if isinstance(period_result, dict):
                    value = period_result.get(stat)
                values.append('N/A' if value is None else f'{value:.6f}')
            print(f'{str(period):<5} ' + "   ".join(values))


def update_model_comparison_workbook(rmse_table):
    workbook_path = base_dir / 'Model_Comparisons.xlsx'
    wb = load_workbook(workbook_path)
    ws = wb['Sheet1']

    excel_to_proj = {
        'ATC': 'ATC',
        'THE BAT': 'THE_BAT',
        'THE BAT X': 'THE_BAT_X',
        'Steamer': 'Steamer',
        'ZIPS': 'ZIPS',
        'ZIPS DC': 'ZIPS_DC',
        'DC': 'DC'
    }

    title_to_stat = {
        'H': 'H',
        '1B%': '1B%',
        '2B%': '2B%',
        '3B%': '3B%',
        'HR% (/PA)': 'HR%',
        'R': 'R',
        'RBI': 'RBI',
        'BB': 'BB',
        'IBB': 'IBB',
        'SO': 'SO',
        'HBP': 'HBP',
        'SF': 'SF',
        'SH': 'SH',
        'SB': 'SB',
        'CS': 'CS',
        'NSB': 'NSB',
        'AVG': 'AVG',
        'OBP': 'OBP',
        'SLG': 'SLG',
        'OPS': 'OPS',
        'BABIP': 'BABIP',
        'K%': 'K%',
        'BB%': 'BB%'
    }

    section_starts = [r for r in range(1, ws.max_row + 1) if ws.cell(r, 1).value in {'H', '3B%', 'RBI', 'SO', 'SH', 'NSB', 'SLG', 'K%'}]
    for start_row in sorted(section_starts, reverse=True):
        if ws.cell(start_row + 3, 1).value != 2025:
            ws.insert_rows(start_row + 3)

    for row in range(1, ws.max_row + 1):
        for col in range(1, ws.max_column + 1):
            title_val = ws.cell(row, col).value
            if title_val is None:
                continue

            stat_key = title_to_stat.get(str(title_val).strip())
            if stat_key is None:
                continue

            header_row = row + 1
            label_rows = {
                'combined': row + 2,
                2025: row + 3,
                2024: row + 4,
                2023: row + 5
            }

            ws.cell(label_rows['combined'], col).value = 'Total'
            ws.cell(label_rows[2025], col).value = 2025
            ws.cell(label_rows[2024], col).value = 2024
            ws.cell(label_rows[2023], col).value = 2023

            for c in range(col + 1, min(col + 8, ws.max_column + 1)):
                excel_model_name = ws.cell(header_row, c).value
                if excel_model_name is None:
                    continue
                proj_key = excel_to_proj.get(str(excel_model_name).strip())
                if proj_key is None:
                    continue

                for period_key, r in label_rows.items():
                    value = rmse_table.get(proj_key, {}).get(period_key, {}).get(stat_key)
                    ws.cell(r, c).value = value

    wb.save(workbook_path)
    print(f'Updated workbook: {workbook_path}')


# rmse_table = {}

# for proj in projections:
#     print(proj)
#     model_dir = base_dir / proj
#     year_paths = {
#         2023: model_dir / f'2023_{proj}_Projections.csv',
#         2024: model_dir / f'2024_{proj}_Projections.csv',
#         2025: model_dir / f'2025_{proj}_Projections.csv'
#     }
#     missing_paths = [str(p) for p in year_paths.values() if not p.exists()]
#     if missing_paths:
#         print(f'  Skipping {proj}; missing files:')
#         for p in missing_paths:
#             print(f'    {p}')
#         print()
#         continue

#     df_2023 = add_derived_stat_columns(pd.read_csv(year_paths[2023]))
#     df_2024 = add_derived_stat_columns(pd.read_csv(year_paths[2024]))
#     df_2025 = add_derived_stat_columns(pd.read_csv(year_paths[2025]))
#     df_3yr = add_derived_stat_columns(pd.concat([df_2023, df_2024, df_2025], ignore_index=True))

#     print_rmse_for_stats(df_2023, '2023')
#     print_rmse_for_stats(df_2024, '2024')
#     print_rmse_for_stats(df_2025, '2025')
#     print_rmse_for_stats(df_3yr, '2023-2025 Combined')
#     print()

#     rmse_table[proj] = {
#         2023: compute_rmse_by_stat(df_2023, rmse_stats),
#         2024: compute_rmse_by_stat(df_2024, rmse_stats),
#         2025: compute_rmse_by_stat(df_2025, rmse_stats),
#         'combined': compute_rmse_by_stat(df_3yr, rmse_stats)
#     }

# update_model_comparison_workbook(rmse_table)

rmse_table = collect_percentage_rmse_2023_2025(
    years=(2023, 2024, 2025),
    projection_systems=projections
)
print_percentage_rmse_2023_2025(
    rmse_table,
    years=(2023, 2024, 2025),
    projection_systems=projections
)


# rand_state = [1,11,21,31,41,51]
# for num in rand_state:
#     # train_rmse, test_rmse, train_r2, test_r2 = polyRegression(path_df, stat_name, num)
#     train_rmse, test_rmse, train_r2, test_r2 = xgboostRegressor(path_df, stat_name, num)
#     # train_rmse, test_rmse, train_r2, test_r2 = elasticNet(path_df, stat_name, num)
#     print(f'Random state: {num}')
#     print(test_rmse)

# print(find_player_name('guerrero jr.','vladimir'))
# é
